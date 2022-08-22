# importing requests package
import requests
import pandas as pd  

# ESG news api
# following query parameters are used
# source, sortBy and apiKey
url = ('https://newsapi.org/v2/everything?'
       'q=ESG SEC&'
       'language=en&'
       'from=2022-08-20&'
       'sortBy=relevancy&'
       'apiKey=bcc83430e9964c40ab973b1012bd82cb')

    
# fetching data in json format
response = requests.get(url)
open_ESG_page = response.json()

display(open_ESG_page)


# getting all articles in a string article
article = open_ESG_page["articles"]

# empty list which will
# contain all trending news
results = []

for ar in article:
    results.append(ar["publishedAt"])
    results.append(ar["title"])
    results.append(ar["description"])
    results.append(ar["source"])
    results.append(ar["url"])

for i in range(len(results)):

# printing all trending news
    print(i + 1, results[i])

# Convert print results to pandas dataframe
df = pd.DataFrame({'columname':results})

# Print dataframe
print(df)

# Transpose every 5 rows into separate column values and save as new dataframe
df2=(pd.DataFrame(df.values.reshape(-1, 5), 
                    columns=['Published Date','Title','Description','Source','URL']))

# Display results for new 2nd dataframe                     
display(df2)

# Convert column from object to string and parse out publisher title from string
df2['Source'] = df2['Source'].astype('string')
df2['Source'] = df2['Source'].str.split("'name':").str[-1]
df2['Source'] = df2['Source'].str.split("'").str[1]

# Display results for updated 2nd dataframe
display(df2)

# Convert column from object to datetime to date
df2['Published Date'] = pd.to_datetime(df2['Published Date'])
df2['Published Date'] = pd.to_datetime(df2['Published Date']).dt.date

# Display results for updated 2nd dataframe
display(df2)

# Activate hyperlink to view articles in URL column and save as new 3rd dataframe
from IPython.display import HTML
df3=HTML(df2.to_html(render_links=True, escape=False))

# Display results for new 3rd dataframe
display(df3)

# Export 3rd dataframe to Excel
df3.to_excel("./", index=False) # enter destination file path

# import openpyxl module
import openpyxl
from openpyxl.styles import borders
from openpyxl.styles.borders import Border

# Give the location of the file
path = "./" # enter source file path

# Assign object values for active workbook and worksheet
wb_obj = openpyxl.load_workbook(path.strip())
sheet_obj = wb_obj.active

# Assign abbreviated variables for workbook and worksheet objects
# Modify column widths
ws = wb_obj
sheet = sheet_obj
sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 75
sheet.column_dimensions['C'].width = 50
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 100

# Save changes to Excel file
ws.save(path)

# Import Patternfill from openpyxl.styles module
from openpyxl.styles import PatternFill

# Assign abbreviated variables for workbook and worksheet objects
wb = wb_obj
ws = sheet_obj

# Format column headers
ws["A1"].fill = PatternFill("solid", start_color="B2B2B2")
ws["B1"].fill = PatternFill("solid", start_color="B2B2B2")
ws["C1"].fill = PatternFill("solid", start_color="B2B2B2")
ws["D1"].fill = PatternFill("solid", start_color="B2B2B2")
ws["E1"].fill = PatternFill("solid", start_color="B2B2B2")
ws["F1"].fill = PatternFill("solid", start_color="B2B2B2")
ws["G1"].fill = PatternFill("solid", start_color="B2B2B2")

# Save changes to Excel file
wb.save(path)
